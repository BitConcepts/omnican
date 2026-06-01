"""Extract J1939DA column headers, all PGNs, and diagnostic DM messages."""
import openpyxl
from pathlib import Path
import json

DA_PATH = Path(r"C:\Users\trist\OneDrive\Desktop\J1939DA_202412\J1939DA DEC2024 r1.xlsx")

print("Loading J1939DA DEC2024 ...")
wb = openpyxl.load_workbook(DA_PATH, read_only=True, data_only=True)
ws = wb["SPs & PGs"]

# Row 4 is the header row
headers = []
for row in ws.iter_rows(min_row=4, max_row=4, values_only=True):
    headers = [str(h) if h is not None else "" for h in row]
    break

print(f"\nSPs & PGs columns ({len(headers)}):")
for i, h in enumerate(headers):
    print(f"  [{i}] {h[:80]}")

# Find key column indices
def col(name_fragment):
    for i, h in enumerate(headers):
        if name_fragment.lower() in h.lower():
            return i
    return -1

pgn_col = col("PGN")
pg_label_col = col("PG Label")
pg_acro_col = col("PG Acronym")
pg_desc_col = col("PG Description")
pdu_col = col("PDU Type")
dl_col = col("PG Data Minimum Length")
tx_rate_col = col("Transmission Rate")
spn_col = col("SPN")
sp_name_col = col("SP Name")
sp_desc_col = col("SP Description")
sp_pos_col = col("SP Start")
sp_len_col = col("SP Length")

print(f"\nKey column indices:")
print(f"  PGN={pgn_col}, PG Label={pg_label_col}, PG Acronym={pg_acro_col}")
print(f"  PDU Type={pdu_col}, Min Length={dl_col}, TX Rate={tx_rate_col}")
print(f"  SPN={spn_col}, SP Name={sp_name_col}, SP Len={sp_len_col}")

# Collect all unique PGNs with their labels
print("\n--- Extracting all PGNs ---")
pgns = {}  # pgn -> {label, acro, desc, pdu}
dm_msgs = {}  # diagnostic messages
spn_map = {}  # spn -> {name, pgn}

row_count = 0
for row in ws.iter_rows(min_row=5, values_only=True):
    if not any(c is not None for c in row):
        continue
    row_count += 1

    pgn_val = row[pgn_col] if pgn_col >= 0 and pgn_col < len(row) else None
    label = row[pg_label_col] if pg_label_col >= 0 and pg_label_col < len(row) else None
    acro = row[pg_acro_col] if pg_acro_col >= 0 and pg_acro_col < len(row) else None
    pdu = row[pdu_col] if pdu_col >= 0 and pdu_col < len(row) else None
    spn_val = row[spn_col] if spn_col >= 0 and spn_col < len(row) else None
    sp_name = row[sp_name_col] if sp_name_col >= 0 and sp_name_col < len(row) else None

    if pgn_val is not None and isinstance(pgn_val, (int, float)):
        pgn_int = int(pgn_val)
        if pgn_int not in pgns:
            pgns[pgn_int] = {
                "label": str(label) if label else "",
                "acro": str(acro) if acro else "",
                "pdu": str(pdu) if pdu else "",
            }
        # Diagnostic messages (DM prefix in acronym or label)
        acro_s = str(acro) if acro else ""
        label_s = str(label) if label else ""
        if acro_s.startswith("DM") or "Diagnostic" in label_s:
            dm_msgs[pgn_int] = {
                "label": label_s,
                "acro": acro_s,
                "pdu": str(pdu) if pdu else "",
            }

    if spn_val is not None and isinstance(spn_val, (int, float)):
        spn_int = int(spn_val)
        if spn_int not in spn_map and sp_name:
            spn_map[spn_int] = {
                "name": str(sp_name),
                "pgn": int(pgn_val) if isinstance(pgn_val, (int, float)) else None,
            }

print(f"\nTotal data rows processed: {row_count}")
print(f"Unique PGNs: {len(pgns)}")
print(f"Diagnostic messages (DM*): {len(dm_msgs)}")
print(f"Unique SPNs sampled: {len(spn_map)}")

# Print all DM messages sorted by PGN
print("\n--- DIAGNOSTIC MESSAGES (DM*) sorted by PGN ---")
for pgn in sorted(dm_msgs.keys()):
    d = dm_msgs[pgn]
    pgn_hex = f"0x{pgn:05X}"
    print(f"  PGN {pgn:6d} ({pgn_hex})  {d['acro']:12s}  {d['label'][:60]}")

# Print all PGNs by range
print("\n--- ALL PGNs by range ---")
sorted_pgns = sorted(pgns.keys())
print(f"  Min PGN: {sorted_pgns[0]} (0x{sorted_pgns[0]:05X})")
print(f"  Max PGN: {sorted_pgns[-1]} (0x{sorted_pgns[-1]:05X})")
print(f"  Total unique PGNs: {len(sorted_pgns)}")

# Sample some important well-known PGNs
important = [0, 256, 512, 61184, 61440, 65226, 65227, 65228, 65229, 65230,
             65231, 65232, 65233, 65234, 65235, 65236, 65237, 65238, 65239,
             65240, 65241, 65242, 65243, 65244, 65245, 65246, 65247, 65248,
             65249, 65250, 65251, 65252, 65253, 65254, 65255, 65256, 65257,
             65258, 65259, 65260, 65261, 65262, 65263, 65264, 65265, 65266,
             60928, 59392, 59904, 60160]
print("\n--- KEY WELL-KNOWN PGNs ---")
for pgn in important:
    if pgn in pgns:
        d = pgns[pgn]
        print(f"  PGN {pgn:6d} (0x{pgn:05X}) {d['acro']:12s} {d['label'][:50]}")

# Save summary
out = {
    "total_pgns": len(pgns),
    "total_dm_msgs": len(dm_msgs),
    "total_spns": len(spn_map),
    "dm_msgs": {str(k): v for k, v in sorted(dm_msgs.items())},
    "all_pgns_sorted": [
        {"pgn": p, "hex": f"0x{p:05X}", "acro": pgns[p]["acro"], "label": pgns[p]["label"]}
        for p in sorted_pgns
    ],
}
out_path = Path(r"C:\Users\trist\Development\BitConcepts\omnican\.specsmith\j1939da_summary.json")
out_path.write_text(json.dumps(out, indent=2), encoding="utf-8")
print(f"\nSummary saved to {out_path}")

wb.close()
print("Done.")
