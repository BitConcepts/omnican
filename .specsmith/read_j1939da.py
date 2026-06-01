"""Extract structure and key rows from J1939DA DEC2024 Excel file."""
import openpyxl
from pathlib import Path

DA_PATH = Path(r"C:\Users\trist\OneDrive\Desktop\J1939DA_202412\J1939DA DEC2024 r1.xlsx")

print(f"Loading {DA_PATH.name} ...")
wb = openpyxl.load_workbook(DA_PATH, read_only=True, data_only=True)
print(f"\nSheets ({len(wb.sheetnames)}):")
for s in wb.sheetnames:
    print(f"  {s}")

SHEETS_TO_SAMPLE = wb.sheetnames  # sample all

for sheet_name in SHEETS_TO_SAMPLE:
    ws = wb[sheet_name]
    print(f"\n{'='*60}")
    print(f"SHEET: {sheet_name}")
    #print(f"  Dimensions: {ws.dimensions}")
    rows = list(ws.iter_rows(max_row=6, values_only=True))
    for i, row in enumerate(rows):
        non_empty = [c for c in row if c is not None]
        if non_empty:
            print(f"  Row {i+1}: {non_empty[:12]}")
    # Count non-empty rows
    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if any(c is not None for c in row):
            count += 1
    print(f"  Data rows (approx): {count}")

wb.close()
print("\nDone.")
